import json
import pandas as pd
from dbimporter.logger import logger

# pip3 install openpyxl
# pip3 install openpyxl-image-loader

import openpyxl
from openpyxl_image_loader import SheetImageLoader

def read_data(filename: str,
              testing: bool = False):

    """
    Read from excel file to check and return dict of sheetname: sheet_data


    Parameters
    ----------

        filename: str
            Path to excel file to be checked for importing

    Returns
    -------

        sheet_names : list
            A list of the sheet names from the file being read
        sheets : dict
            A dictionary of the sheet_names and the pandas dataframe parsed from those sheets
    """

    if filename.lower().endswith('.csv'):
        df = pd.read_csv(filename, on_bad_lines='skip')
    elif filename.lower().endswith('.xlsx'):
        df = pd.ExcelFile(filename)

    try:
        #check_sheets(df.sheet_names, self.expected_json)
        sheets = {}
        sheet_names = df.sheet_names
        for i in sheet_names:
            sheets[i] = df.parse(i)
    except AttributeError:
        logger.info("File format has no sheets")
        sheet_names = [filename]
        sheets = {filename: df}

    return sheet_names, sheets

def get_image_test(filename):

    sheetnames, sheets = read_data(filename)

    #loading the Excel File and the sheet
    pxl_doc = openpyxl.load_workbook(filename)
    sheet = pxl_doc[sheetnames[0]]

    #calling the image_loader
    image_loader = SheetImageLoader(sheet)

    print(sheet._images)

    #get the image (put the cell you need instead of 'A1')
    # image = image_loader.get('A1')

    # #showing the image
    # image.show()

    # #saving the image
    # image.save('my_path/image_name.jpg')